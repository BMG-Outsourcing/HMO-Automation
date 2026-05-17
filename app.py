"""
BMG-HMO Billing File Processor — Streamlit App  v10.8.2

FIXES over v10.8.1:
  ✓ FIX 1 — Removed wb.close() calls (openpyxl has no .close())
  ✓ FIX 2 — BytesIO buffers no longer closed while workbook still alive
  ✓ FIX 3 — prog.empty() now called on success to prevent white-flash rerun loop
  ✓ FIX 4 — Removed value=False from checkbox to stop rerun loop
  ✓ FIX 5 — Fixed wb.worksheets.index(ws) (was broken wb.index(ws))
  ✓ FIX 6 — Fixed empty label warnings: file_uploader and text_input now have
             non-empty labels with label_visibility="collapsed"
"""

import os, io, json, re, gc, time
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

# ─── PATHS ────────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
MASTER_PATH = os.path.join(BASE_DIR, "master_reference.json")
LOGO_PATH   = os.path.join(BASE_DIR, "images", "logo.png")

# ─── CONSTANTS ────────────────────────────────────────────────────────────────
HEADER_ROW = 18

FINANCIAL_COLS = ["Medical", "VAT", "Total Membership Fee"]

SC_TO_CAT = {
    "HMO Incidentals - SO":          "COS",
    "HMO SO Advances":               "Advances",
    "HMO - SO Advances":             "Advances",
    "Medical Expense":               "Medical Expense",
    "HMO Internal Advances":         "Advances",
    "HMO - Internal Advances":       "Advances",
    "Medical Expense HMO - Drivers": "Medical Expense",
}

SC_SORT_ORDER = [
    "HMO Incidentals - SO",
    "HMO SO Advances",
    "HMO - SO Advances",
    "HMO Internal Advances",
    "HMO - Internal Advances",
    "Medical Expense",
    "Medical Expense HMO - Drivers",
]

SC_AUTO_LABEL = "(auto-derive from Entity)"

WF_ONLY_COLS = {"Entity", "System Coding", "Reference"}

KEEP_SHEETS = {"Attachment", "Working File"}

WF_OUTPUT_COLS = [
    "Id Number",
    "Name",
    "Entity",
    "System Coding",
    "Medical",
    "VAT",
    "Total Membership Fee",
]

WF_READ_COLS = WF_OUTPUT_COLS + ["Reference"]

_ID_PATTERN = re.compile(r'^\d{4}-')

_HEADER_ID_VALUES = {"nan", "none", "", "id number", "amount in words", "#"}

HEADER_COLOR    = "C00000"
HIGHLIGHT_COLOR = "FFF2CC"

# ─── Entity → output sheet name ───────────────────────────────────────────────
def _entity_sheet(entity: str) -> str:
    el = entity.strip().lower()
    if "sleek" in el:
        return "Sleek"
    if "nyfd" in el:
        return "NYFD"
    if "bmg internal" in el or "bmgo internal" in el:
        return "BMG Internal"
    return "SO"

ENTITY_SHEETS = ["Sleek", "NYFD", "BMG Internal", "SO"]

# ─── Sleek sub-groups ─────────────────────────────────────────────────────────
def _sleek_subgroup(entity: str) -> str:
    el = entity.strip().lower()
    if "sg" in el and ("r&d" in el or "r and d" in el):
        return "Sleek - SG (R&D)"
    if "sg" in el:
        return "Sleek - SG"
    if "hk" in el and "cpa" in el:
        return "Sleek - HK (CPA)"
    if "hk" in el:
        return "Sleek - HK"
    if "uk" in el:
        return "Sleek - UK"
    if "au" in el:
        return "Sleek - AU"
    return "Sleek - SG"

SLEEK_SUBGROUPS = [
    "Sleek - SG",
    "Sleek - HK",
    "Sleek - HK (CPA)",
    "Sleek - SG (R&D)",
    "Sleek - UK",
    "Sleek - AU",
]

SECTION_HEADER_COLOR = "404040"
SECTION_HEADER_BG    = "D9D9D9"


# ─── Style factories ──────────────────────────────────────────────────────────
def _make_header_fill():
    return PatternFill("solid", fgColor="C00000")

def _make_header_font():
    return Font(bold=True, color="FFFFFF")

def _make_highlight_fill():
    return PatternFill("solid", fgColor="FFF2CC")

def _make_no_fill():
    return PatternFill(fill_type=None)


# ─── SAFE HELPERS ─────────────────────────────────────────────────────────────
def _s(val) -> str:
    if isinstance(val, pd.Series):
        val = val.iloc[0] if not val.empty else ""
    if val is None:
        return ""
    try:
        if isinstance(val, float) and pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    return str(val).strip()


def _isna(val) -> bool:
    if isinstance(val, pd.Series):
        return val.empty or bool(val.isna().all())
    try:
        return bool(pd.isna(val))
    except (TypeError, ValueError):
        return False


def _to_numeric(val):
    if _isna(val) or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return val


def _concat_preserve(base: pd.DataFrame, *others) -> pd.DataFrame:
    saved_attrs = dict(base.attrs)
    result = pd.concat([base, *others], ignore_index=True)
    result.attrs.update(saved_attrs)
    return result


# ─── ROW FILTER ───────────────────────────────────────────────────────────────
def _is_header_id(val) -> bool:
    cleaned = str(val).strip().lower()
    return cleaned in _HEADER_ID_VALUES


def _is_entirely_blank_row(row: pd.Series) -> bool:
    for v in row:
        s = str(v).strip() if v is not None else ""
        if s and s.lower() not in ("nan", "none"):
            return False
    return True


def _valid_rows(df: pd.DataFrame) -> pd.Series:
    return ~df.apply(_is_entirely_blank_row, axis=1)


def _log_id_info(raw_series: pd.Series, label: str, logs: list):
    cleaned = raw_series.astype(str).str.strip()

    header_like = cleaned[raw_series.apply(_is_header_id)]
    header_like = header_like[header_like.str.lower() != "nan"]
    if not header_like.empty:
        logs.append(
            f"  ℹ {label}: {len(header_like)} header-like ID value(s) found "
            f"(kept if row has other data, dropped only if fully blank):"
        )
        for val in header_like.unique()[:10]:
            logs.append(f"    ↳ '{val}'")

    non_std = cleaned[
        ~raw_series.apply(_is_header_id) &
        ~cleaned.str.match(r'^\d{4}-')
    ]
    if not non_std.empty:
        logs.append(
            f"  ℹ {label}: {len(non_std)} row(s) have non-standard ID format (kept):"
        )
        for val in non_std.unique()[:10]:
            logs.append(f"    ↳ '{val}'")


# ─── CLASSIFICATION ───────────────────────────────────────────────────────────
def derive_system_coding(entity: str) -> str:
    e = _s(entity)
    if not e:
        return ""
    el = e.lower()
    eu = e.upper()
    if "BMG INTERNAL" in eu or "BMGO INTERNAL" in eu:
        return "HMO Internal Advances" if "advance" in el else "Medical Expense"
    if "advance" in el:
        return "HMO SO Advances"
    return "HMO Incidentals - SO"


def get_company_name(entity: str) -> str:
    e = _s(entity)
    e = re.sub(r'\s*-\s*[Dd]epend.*$', '', e)
    e = re.sub(r'\s*-\s*[Aa]dvance.*$', '', e)
    return e.strip()


# ─── SHEET READER ─────────────────────────────────────────────────────────────
def _find_header_row(ws, fallback: int = HEADER_ROW) -> int:
    for r in range(1, 31):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and "id number" in str(v).strip().lower():
                return r
    return fallback


def _read_sheet(wb, sheet_name: str) -> pd.DataFrame:
    ws = wb[sheet_name]
    hdr_row = _find_header_row(ws, fallback=HEADER_ROW)
    raw_hdr = [ws.cell(row=hdr_row, column=c).value
               for c in range(1, ws.max_column + 1)]
    data = []
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if any(v is not None for v in row):
            data.append(list(row))

    cols, seen = [], {}
    for h in raw_hdr:
        key = _s(h).strip() if h is not None else "_blank"
        if not key:
            key = "_blank"
        if key in seen:
            seen[key] += 1
            cols.append(f"{key}.{seen[key]}")
        else:
            seen[key] = 0
            cols.append(key)

    df = pd.DataFrame(data, columns=cols)

    if sheet_name == "Working File":
        keep = [c for c in df.columns if c in WF_READ_COLS]
        df = df[keep].copy()

    return df


# ─── PERIOD LABEL DETECTOR ────────────────────────────────────────────────────
def _detect_period_label(wb) -> str:
    month_pattern = re.compile(r'month', re.IGNORECASE)
    for sname in wb.sheetnames:
        ws = wb[sname]
        for r in range(1, 21):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v and month_pattern.search(str(v)):
                    return str(v).strip()
    return ""


# ─── PHASE 1: VALIDATION ──────────────────────────────────────────────────────
def phase1_validate(wb, filename: str) -> dict:
    result = {
        "monthly":          False,
        "has_attachment":   "Attachment"   in wb.sheetnames,
        "has_working_file": "Working File" in wb.sheetnames,
    }
    if "MONTHLY" in filename.upper():
        result["monthly"] = True
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            for cell in row:
                if cell and "MONTHLY" in str(cell).upper():
                    result["monthly"] = True
    if not result["monthly"]:
        raise ValueError("Validation failed: 'MONTHLY' keyword not found.")
    if not result["has_working_file"]:
        raise ValueError("'Working File' sheet not found in workbook.")
    return result


# ─── PHASE 2: SYNC ────────────────────────────────────────────────────────────
def phase2_sync(wb) -> tuple[pd.DataFrame, list[str]]:
    logs: list[str] = []
    gc.collect()

    wf = _read_sheet(wb, "Working File")
    logs.append(f"  Working File columns found    : {list(wf.columns)}")

    if "Id Number" not in wf.columns:
        id_col_candidates = [c for c in wf.columns if "id" in c.lower() and "number" in c.lower()]
        if id_col_candidates:
            wf = wf.rename(columns={id_col_candidates[0]: "Id Number"})
            logs.append(f"  ⚠ Renamed '{id_col_candidates[0]}' → 'Id Number' in Working File")
        else:
            raise KeyError("'Id Number' column not found in Working File sheet.")

    wf["Id Number"] = wf["Id Number"].astype(str).str.strip()
    _log_id_info(wf["Id Number"], "Working File", logs)

    wf_total_before = len(wf)
    valid_mask = _valid_rows(wf)
    wf_dropped = wf_total_before - valid_mask.sum()
    wf = wf[valid_mask].copy()

    for col in ("Entity", "System Coding", "Reference"):
        if col not in wf.columns:
            wf[col] = ""

    logs.append(f"  Working File rows total       : {wf_total_before}")
    logs.append(f"  Working File rows (kept)      : {len(wf)}")
    if wf_dropped:
        logs.append(f"  Working File rows dropped     : {wf_dropped} (fully blank rows only)")

    wf_dup_mask = wf.duplicated(subset="Id Number", keep=False)
    if wf_dup_mask.any():
        dup_ids = wf.loc[wf_dup_mask, "Id Number"].unique()
        logs.append(f"  ⚠ Duplicate IDs in Working File ({len(dup_ids)}) — ALL rows kept:")
        for rid in dup_ids[:10]:
            logs.append(f"    ↳ {rid}")
        if len(dup_ids) > 10:
            logs.append(f"    … and {len(dup_ids) - 10} more")

    if "Attachment" not in wb.sheetnames:
        logs.append("  ⚠ No Attachment sheet — Working File used as-is.")
        wf = _finalise_sc(wf)
        for col in WF_OUTPUT_COLS:
            if col not in wf.columns:
                wf[col] = ""
        wf.attrs["att_ids"] = set()
        return wf, logs

    att = _read_sheet(wb, "Attachment")
    logs.append(
        f"  Attachment columns found      : "
        f"{list(att.columns)[:12]}{'...' if len(att.columns) > 12 else ''}"
    )

    if "Id Number" not in att.columns:
        id_col_candidates = [c for c in att.columns if "id" in c.lower() and "number" in c.lower()]
        if id_col_candidates:
            att = att.rename(columns={id_col_candidates[0]: "Id Number"})
            logs.append(f"  ⚠ Renamed '{id_col_candidates[0]}' → 'Id Number' in Attachment")
        else:
            logs.append("  ✗ CRITICAL: 'Id Number' column not found in Attachment — no employees synced!")
            wf.attrs["att_ids"] = set()
            return wf, logs

    att["Id Number"] = att["Id Number"].astype(str).str.strip()
    _log_id_info(att["Id Number"], "Attachment", logs)

    att_total_before = len(att)
    att_valid_mask = _valid_rows(att)
    att_dropped = att_total_before - att_valid_mask.sum()
    att = att[att_valid_mask].copy()

    logs.append(f"  Attachment rows total         : {att_total_before}")
    logs.append(f"  Attachment rows (kept)        : {len(att)}")
    if att_dropped:
        logs.append(f"  Attachment rows dropped       : {att_dropped} (fully blank rows only)")

    att_dup_mask = att.duplicated(subset="Id Number", keep=False)
    if att_dup_mask.any():
        dup_ids = att.loc[att_dup_mask, "Id Number"].unique()
        logs.append(
            f"  ⚠ Duplicate IDs in Attachment ({len(dup_ids)}) — "
            f"all rows tracked; last row used for financial refresh:"
        )
        for rid in dup_ids[:10]:
            logs.append(f"    ↳ {rid}")
        if len(dup_ids) > 10:
            logs.append(f"    … and {len(dup_ids) - 10} more")

    att_ids = set(att["Id Number"].tolist())
    att_lkp = (
        att.drop_duplicates(subset="Id Number", keep="last")
           .set_index("Id Number")
           .to_dict("index")
    )
    wf_ids = set(wf["Id Number"].tolist())
    wf.attrs["att_ids"] = att_ids

    for col in WF_OUTPUT_COLS:
        if col not in wf.columns:
            wf[col] = ""

    refreshed      = 0
    discrepancies: list[str] = []
    for idx in wf.index:
        rid = _s(wf.at[idx, "Id Number"])
        if rid not in att_lkp:
            continue
        att_row = att_lkp[rid]
        changed = False
        for fc in FINANCIAL_COLS:
            nv = att_row.get(fc)
            if not _isna(nv):
                old_val = wf.at[idx, fc] if fc in wf.columns else ""
                if _s(old_val) != _s(nv):
                    wf.at[idx, fc] = _to_numeric(nv)
                    changed = True
        if changed:
            refreshed += 1
        for dc in ("Name",):
            wv = _s(wf.at[idx, dc] if dc in wf.columns else "")
            av = _s(att_row.get(dc, ""))
            if wv and av and wv != av:
                discrepancies.append(f"    {rid} | {dc}: WF='{wv}'  ATT='{av}'")

    logs.append(f"  Financial cols refreshed      : {refreshed} records")
    if discrepancies:
        logs.append(f"  ⚠ Name mismatches ({len(discrepancies)}):")
        logs.extend(discrepancies[:15])
        if len(discrepancies) > 15:
            logs.append(f"    … and {len(discrepancies) - 15} more")

    master = load_master()
    new_count = 0
    for rid in sorted(att_ids - wf_ids):
        att_row      = att_lkp[rid]
        master_entry = master.get(rid, {})
        entity = _s(master_entry.get("Entity", ""))
        if entity:
            sc = derive_system_coding(entity)
            if not sc:
                sc = _s(master_entry.get("SystemCoding", ""))
        else:
            sc = _s(master_entry.get("SystemCoding", ""))

        nr: dict = {col: "" for col in WF_OUTPUT_COLS}
        nr["Id Number"]     = rid
        nr["Name"]          = _s(att_row.get("Name", ""))
        nr["Entity"]        = entity
        nr["System Coding"] = sc
        for fc in FINANCIAL_COLS:
            val    = att_row.get(fc)
            nr[fc] = None if _isna(val) else _to_numeric(val)

        wf = _concat_preserve(wf, pd.DataFrame([nr]))
        new_count += 1
        sc_display = sc if sc else "(blank — fill manually)"
        logs.append(
            f"  + New from Attachment         : {rid}  {_s(att_row.get('Name', ''))}  "
            f"[SC: {sc_display}]"
        )

    if new_count:
        logs.append(f"  Total new rows synced         : {new_count}")

    wf_only = wf_ids - att_ids
    if wf_only:
        logs.append(f"  ⚠ WF-only / manual ({len(wf_only)}) — kept, highlighted yellow:")
        for rid in sorted(wf_only)[:20]:
            nm = wf.loc[wf["Id Number"] == rid, "Name"]
            logs.append(f"    ↳ {rid}  {_s(nm.values[0]) if len(nm) else ''}")
        if len(wf_only) > 20:
            logs.append(f"    … and {len(wf_only) - 20} more")

    wf = _finalise_sc(wf)
    logs.append(f"  Final Working File rows       : {len(wf)}")

    logs.append("")
    logs.append("  Entity Segregation Preview:")
    for sname in ENTITY_SHEETS:
        count = sum(
            1 for _, row in wf.iterrows()
            if _entity_sheet(_s(row.get("Entity", ""))) == sname
        )
        logs.append(f"    → {sname:<14}: {count} row(s)")

    gc.collect()
    return wf, logs


def _finalise_sc(df: pd.DataFrame) -> pd.DataFrame:
    saved_attrs = dict(df.attrs)
    if "System Coding" not in df.columns:
        df["System Coding"] = ""
    if "Entity" not in df.columns:
        df["Entity"] = ""

    def _sc(row):
        existing = _s(row.get("System Coding", ""))
        if existing:
            return existing
        entity = _s(row.get("Entity", ""))
        if not entity:
            return ""
        return derive_system_coding(entity)

    df["System Coding"] = df.apply(_sc, axis=1)
    df.attrs.update(saved_attrs)
    return df


# ─── MASTER REFERENCE ─────────────────────────────────────────────────────────
def load_master() -> dict:
    if os.path.exists(MASTER_PATH):
        with open(MASTER_PATH) as f:
            return json.load(f)
    return {}


def save_master(master: dict):
    with open(MASTER_PATH, "w") as f:
        json.dump(master, f, indent=2)


def update_master(df: pd.DataFrame, master: dict) -> tuple[dict, dict]:
    log  = {"added": [], "updated": [], "missing": []}
    live = set(df["Id Number"].astype(str).str.strip())
    for _, row in df.iterrows():
        rid    = _s(row.get("Id Number", ""))
        entity = _s(row.get("Entity", ""))
        sc     = _s(row.get("System Coding", ""))
        if not rid:
            continue
        entry = {
            "Name":         _s(row.get("Name", "")),
            "Entity":       entity,
            "Company":      get_company_name(entity),
            "SystemCoding": sc,
            "Category":     SC_TO_CAT.get(sc, "COS"),
        }
        if rid not in master:
            master[rid] = entry
            log["added"].append(rid)
        else:
            for k, v in entry.items():
                if v or k not in master[rid]:
                    master[rid][k] = v
            log["updated"].append(rid)
    for mid in list(master.keys()):
        if mid not in live:
            master[mid]["status"] = "missing_from_upload"
            log["missing"].append(mid)
        else:
            master[mid].pop("status", None)
    return master, log


# ─── OPENPYXL STYLE HELPERS ───────────────────────────────────────────────────
def _safe_font(src) -> Font:
    if src is None:
        return Font()
    try:
        return Font(
            name=src.name, size=src.size, bold=src.bold, italic=src.italic,
            underline=src.underline, strike=src.strike,
            color=copy(src.color) if src.color else None,
            vertAlign=src.vertAlign, charset=src.charset,
            family=src.family, scheme=src.scheme,
        )
    except Exception:
        return Font()


def _safe_fill(src) -> PatternFill:
    if src is None:
        return PatternFill()
    try:
        return copy(src) if isinstance(src, PatternFill) else PatternFill()
    except Exception:
        return PatternFill()


def _safe_border(src) -> Border:
    if src is None:
        return Border()
    try:
        return copy(src)
    except Exception:
        return Border()


def _safe_alignment(src) -> Alignment:
    if src is None:
        return Alignment()
    try:
        return copy(src)
    except Exception:
        return Alignment()


def _copy_style(src, dst):
    if src.has_style:
        dst.font          = _safe_font(src.font)
        dst.border        = _safe_border(src.border)
        dst.fill          = _safe_fill(src.fill)
        dst.number_format = src.number_format
        dst.protection    = copy(src.protection)
        dst.alignment     = _safe_alignment(src.alignment)


# ─── WORKBOOK SCRUBBER ────────────────────────────────────────────────────────
def _scrub_workbook(wb):
    try:
        wb._external_links.clear()
    except AttributeError:
        pass
    try:
        pkg_rels = wb._part._rels
        to_drop = [k for k, r in pkg_rels.items()
                   if "externalLink" in r.reltype or "externalLink" in str(r.target_ref)]
        for k in to_drop:
            del pkg_rels[k]
    except Exception:
        pass

    for ws in wb.worksheets:
        ws._drawing = None
        try:
            sheet_rels = ws._part._rels
            to_drop = [k for k, r in sheet_rels.items()
                       if "drawing" in r.reltype.lower()]
            for k in to_drop:
                del sheet_rels[k]
        except Exception:
            pass
        try:
            ws.legacy_drawing = None
        except Exception:
            pass

    return wb


# ─── SC SORT KEY ──────────────────────────────────────────────────────────────
def _sc_sort_key(sc_value: str) -> tuple[int, str]:
    sc = _s(sc_value)
    try:
        return (SC_SORT_ORDER.index(sc), sc)
    except ValueError:
        return (len(SC_SORT_ORDER), sc)


def _sort_by_sc(df: pd.DataFrame) -> pd.DataFrame:
    if "System Coding" not in df.columns or df.empty:
        return df
    saved_attrs = dict(df.attrs)
    df = df.copy()
    df["_sc_rank"] = df["System Coding"].apply(lambda v: _sc_sort_key(v)[0])
    df = df.sort_values(["_sc_rank", "System Coding"], kind="stable").drop(
        columns=["_sc_rank"]
    ).reset_index(drop=True)
    df.attrs.update(saved_attrs)
    return df


# ─── SUMMARY SHEET WRITER ─────────────────────────────────────────────────────
def _write_summary_sheet(ws, df: pd.DataFrame, period_label: str = ""):
    NUM_FMT   = "#,##0.00"
    CAL_FONT  = "Calibri"

    GREEN_FILL = PatternFill("solid", fgColor="C00000")
    WHITE_FONT = Font(bold=True, color="FFFFFF", name=CAL_FONT, size=11)
    BOLD_FONT  = Font(bold=True, name=CAL_FONT, size=11)
    REG_FONT   = Font(name=CAL_FONT, size=11)
    TITLE_FONT = Font(bold=True, name=CAL_FONT, size=13)
    MED_SIDE   = Side(style="medium", color="000000")

    def med_tb():
        return Border(top=MED_SIDE, bottom=MED_SIDE)

    def _group_sum(group_name: str, col: str, advance: bool | None = None) -> float:
        mask = df["Entity"].apply(lambda e: _entity_sheet(_s(e))) == group_name
        if advance is True:
            mask &= df.get("System Coding", pd.Series([""] * len(df), index=df.index)) \
                       .astype(str).str.contains("Advance", case=False, na=False)
        elif advance is False:
            mask &= ~df.get("System Coding", pd.Series([""] * len(df), index=df.index)) \
                        .astype(str).str.contains("Advance", case=False, na=False)
        sub = df[mask]
        if col not in sub.columns or sub.empty:
            return 0.0
        return float(pd.to_numeric(sub[col], errors="coerce").fillna(0).sum())

    GROUP_LABELS = [
        ("SO",           "SO"),
        ("BMG Internal", "BMGO Internal"),
        ("Sleek",        "Sleek"),
        ("NYFD",         "NYFD"),
    ]

    data_rows = []
    for group, label in GROUP_LABELS:
        principal = _group_sum(group, "Medical",              advance=False)
        dependent = _group_sum(group, "VAT",                  advance=False)
        advances  = _group_sum(group, "Medical",              advance=True)
        total     = _group_sum(group, "Total Membership Fee", advance=None)
        data_rows.append((label, principal, dependent, advances, total))

    TITLE_TEXTS = [
        "BMG Outsourcing",
        "Schedule of HMO",
        period_label if period_label else "For the Month of ___",
    ]
    for r_idx, txt in enumerate(TITLE_TEXTS, start=1):
        ws.merge_cells(start_row=r_idx, start_column=1,
                       end_row=r_idx,   end_column=5)
        cell           = ws.cell(row=r_idx, column=1, value=txt)
        cell.font      = TITLE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r_idx].height = 20

    ws.row_dimensions[4].height = 8

    HDR_ROW  = 5
    COL_HDRS = ["Entity", "Principal", "Dependent", "Advances", "Total"]
    for ci, h in enumerate(COL_HDRS, start=1):
        cell           = ws.cell(row=HDR_ROW, column=ci, value=h)
        cell.font      = WHITE_FONT
        cell.fill      = GREEN_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[HDR_ROW].height = 18

    DATA_START = HDR_ROW + 1
    for i, (label, principal, dependent, advances, total) in enumerate(data_rows):
        r = DATA_START + i
        cell_a      = ws.cell(row=r, column=1, value=label)
        cell_a.font = REG_FONT
        for ci, val in enumerate([principal, dependent, advances, total], start=2):
            cell               = ws.cell(row=r, column=ci, value=val if val else None)
            cell.number_format = NUM_FMT
            cell.font          = REG_FONT
            cell.alignment     = Alignment(horizontal="right")
        ws.row_dimensions[r].height = 16

    SPACER1   = DATA_START + len(data_rows)
    TOTAL_ROW = SPACER1 + 1
    ws.row_dimensions[SPACER1].height = 6

    cell_t        = ws.cell(row=TOTAL_ROW, column=1, value="Total")
    cell_t.font   = BOLD_FONT
    cell_t.border = med_tb()
    for ci in range(2, 6):
        col_letter = get_column_letter(ci)
        formula    = (f"=SUM({col_letter}{DATA_START}:"
                      f"{col_letter}{DATA_START + len(data_rows) - 1})")
        cell               = ws.cell(row=TOTAL_ROW, column=ci, value=formula)
        cell.number_format = NUM_FMT
        cell.font          = BOLD_FONT
        cell.alignment     = Alignment(horizontal="right")
        cell.border        = med_tb()
    ws.row_dimensions[TOTAL_ROW].height = 18

    SPACER2 = TOTAL_ROW + 1
    INV_ROW = SPACER2  + 1
    ws.row_dimensions[SPACER2].height = 6

    cell_inv      = ws.cell(row=INV_ROW, column=1, value="Invoice")
    cell_inv.font = BOLD_FONT
    inv_e               = ws.cell(row=INV_ROW, column=5, value=None)
    inv_e.number_format = NUM_FMT
    inv_e.font          = BOLD_FONT
    inv_e.alignment     = Alignment(horizontal="right")
    ws.row_dimensions[INV_ROW].height = 16

    SPACER3 = INV_ROW  + 1
    CHK_ROW = SPACER3  + 1
    ws.row_dimensions[SPACER3].height = 6

    cell_chk      = ws.cell(row=CHK_ROW, column=1, value="To Check")
    cell_chk.font = BOLD_FONT
    chk_e               = ws.cell(row=CHK_ROW, column=5,
                                   value=f"=E{TOTAL_ROW}-E{INV_ROW}")
    chk_e.number_format = NUM_FMT
    chk_e.font          = BOLD_FONT
    chk_e.alignment     = Alignment(horizontal="right")
    ws.row_dimensions[CHK_ROW].height = 16

    ws.column_dimensions["A"].width = 22
    for col_letter in ["B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = 17


# ─── WRITE DATA TABLE — Working File ──────────────────────────────────────────
def _write_data_sheet(ws, df_subset: pd.DataFrame, att_ids: set[str]):
    _write_table(ws, df_subset, att_ids, header_row=HEADER_ROW)


# ─── WRITE DATA TABLE — Entity sheets ─────────────────────────────────────────
def _write_data_sheet_entity(ws, df_subset: pd.DataFrame, att_ids: set[str]):
    _write_table(ws, df_subset, att_ids, header_row=1)
    if not df_subset.empty:
        total_row = 1 + 1 + len(df_subset)
        _write_grand_totals_row(ws, df_subset, total_row)


# ─── SHARED TABLE WRITER ──────────────────────────────────────────────────────
def _write_table(ws, df_subset: pd.DataFrame, att_ids: set[str], header_row: int):
    df_reset    = df_subset.reset_index(drop=True)
    n_data      = len(df_reset)
    old_max_row = ws.max_row

    for col_i, col_name in enumerate(WF_OUTPUT_COLS, start=1):
        cell           = ws.cell(row=header_row, column=col_i, value=col_name)
        cell.font      = _make_header_font()
        cell.fill      = _make_header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_i in range(len(WF_OUTPUT_COLS) + 1, ws.max_column + 1):
        cell       = ws.cell(row=header_row, column=col_i)
        cell.value = None
        cell.fill  = _make_no_fill()
        cell.font  = Font()

    for i, (_, row) in enumerate(df_reset.iterrows()):
        write_r   = header_row + 1 + i
        rid       = _s(row.get("Id Number", ""))
        is_manual = bool(att_ids) and rid not in att_ids

        for col_i, col_name in enumerate(WF_OUTPUT_COLS, start=1):
            val = row.get(col_name, None)
            if isinstance(val, pd.Series):
                val = val.iloc[0] if not val.empty else None
            if _isna(val) or val == "":
                val = None
            if col_name in FINANCIAL_COLS and val is not None:
                try:
                    val = round(float(val), 2)
                except (TypeError, ValueError):
                    pass
            cell      = ws.cell(row=write_r, column=col_i, value=val)
            cell.fill = _make_highlight_fill() if is_manual else _make_no_fill()
            if col_name in FINANCIAL_COLS and val is not None:
                cell.number_format = "#,##0.00"

        for col_i in range(len(WF_OUTPUT_COLS) + 1, ws.max_column + 1):
            cell       = ws.cell(row=write_r, column=col_i)
            cell.value = None
            cell.fill  = _make_no_fill()

    first_leftover = header_row + 1 + n_data
    if old_max_row >= first_leftover:
        for r in range(first_leftover, old_max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell       = ws.cell(row=r, column=c)
                cell.value = None
                cell.fill  = _make_no_fill()

    col_widths = [len(h) for h in WF_OUTPUT_COLS]
    for _, row in df_reset.iterrows():
        for ci, col_name in enumerate(WF_OUTPUT_COLS):
            val = row.get(col_name, None)
            if val is not None and not _isna(val) and str(val).strip():
                col_widths[ci] = max(col_widths[ci], len(str(val)))
    for ci, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = min(width + 2, 50)

    last_col_letter = get_column_letter(len(WF_OUTPUT_COLS))
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{header_row + n_data}"
    ws.auto_filter.filterColumn.clear()


# ─── GRAND TOTALS ROW ─────────────────────────────────────────────────────────
def _write_grand_totals_row(ws, df_full: pd.DataFrame, current_row: int):
    NUM_FMT       = "#,##0.00"
    TOTAL_FILL    = PatternFill("solid", fgColor="C00000")
    TOTAL_FONT    = Font(bold=True, color="FFFFFF", size=11)
    TOTAL_ALIGN_L = Alignment(horizontal="left",  vertical="center", indent=1)
    TOTAL_ALIGN_R = Alignment(horizontal="right", vertical="center")
    MED_TOP       = Border(top=Side(style="medium", color="000000"))

    totals: dict[str, float] = {}
    for fc in FINANCIAL_COLS:
        if fc in df_full.columns:
            totals[fc] = round(
                float(pd.to_numeric(df_full[fc], errors="coerce").fillna(0).sum()), 2
            )
        else:
            totals[fc] = 0.0

    for col_i, col_name in enumerate(WF_OUTPUT_COLS, start=1):
        cell        = ws.cell(row=current_row, column=col_i)
        cell.fill   = TOTAL_FILL
        cell.border = MED_TOP

        if col_name == "Id Number":
            cell.value     = "GRAND TOTAL"
            cell.font      = TOTAL_FONT
            cell.alignment = TOTAL_ALIGN_L
        elif col_name in FINANCIAL_COLS:
            cell.value         = totals.get(col_name, 0.0)
            cell.font          = TOTAL_FONT
            cell.alignment     = TOTAL_ALIGN_R
            cell.number_format = NUM_FMT
        else:
            cell.value = None
            cell.font  = TOTAL_FONT

    ws.row_dimensions[current_row].height = 18


# ─── SLEEK SHEET WRITER ───────────────────────────────────────────────────────
def _write_sleek_sheet(ws, df_sleek: pd.DataFrame, att_ids: set[str]):
    n_cols      = len(WF_OUTPUT_COLS)
    last_letter = get_column_letter(n_cols)

    df_work = df_sleek.copy().reset_index(drop=True)
    df_work["_sub"] = df_work["Entity"].apply(lambda e: _sleek_subgroup(_s(e)))

    groups_in_order = []
    for label in SLEEK_SUBGROUPS:
        grp = df_work[df_work["_sub"] == label].drop(columns=["_sub"]).copy()
        if grp.empty:
            continue
        grp = _sort_by_sc(grp)
        if att_ids:
            is_att = grp["Id Number"].isin(att_ids)
            grp = pd.concat([grp[is_att], grp[~is_att]], ignore_index=True)
        groups_in_order.append((label, grp))

    current_row  = 1
    first_filter = None
    col_widths   = [len(h) for h in WF_OUTPUT_COLS]

    section_fill  = PatternFill("solid", fgColor=SECTION_HEADER_BG)
    section_font  = Font(bold=True, color=SECTION_HEADER_COLOR, size=11)
    section_align = Alignment(horizontal="left", vertical="center",
                              indent=1, wrap_text=False)

    for label, grp in groups_in_order:
        sec_cell           = ws.cell(row=current_row, column=1, value=label)
        sec_cell.font      = section_font
        sec_cell.fill      = section_fill
        sec_cell.alignment = section_align
        ws.row_dimensions[current_row].height = 18

        for col_i in range(2, n_cols + 1):
            c       = ws.cell(row=current_row, column=col_i)
            c.fill  = section_fill
            c.value = None
        try:
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row,   end_column=n_cols
            )
        except Exception:
            pass
        current_row += 1

        col_hdr_row = current_row
        for col_i, col_name in enumerate(WF_OUTPUT_COLS, start=1):
            cell           = ws.cell(row=col_hdr_row, column=col_i, value=col_name)
            cell.font      = _make_header_font()
            cell.fill      = _make_header_fill()
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[col_hdr_row].height = 15
        current_row += 1

        if first_filter is None:
            first_filter = col_hdr_row

        for _, row in grp.iterrows():
            rid       = _s(row.get("Id Number", ""))
            is_manual = bool(att_ids) and rid not in att_ids

            for col_i, col_name in enumerate(WF_OUTPUT_COLS, start=1):
                val = row.get(col_name, None)
                if isinstance(val, pd.Series):
                    val = val.iloc[0] if not val.empty else None
                if _isna(val) or val == "":
                    val = None
                if col_name in FINANCIAL_COLS and val is not None:
                    try:
                        val = round(float(val), 2)
                    except (TypeError, ValueError):
                        pass
                cell      = ws.cell(row=current_row, column=col_i, value=val)
                cell.fill = _make_highlight_fill() if is_manual else _make_no_fill()
                if col_name in FINANCIAL_COLS and val is not None:
                    cell.number_format = "#,##0.00"

            for ci, col_name in enumerate(WF_OUTPUT_COLS):
                v = row.get(col_name, None)
                if v is not None and not _isna(v) and str(v).strip():
                    col_widths[ci] = max(col_widths[ci], len(str(v)))

            current_row += 1

        current_row += 1  # spacer between groups

    for ci, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = min(width + 2, 50)

    if first_filter is not None:
        ws.auto_filter.ref = f"A{first_filter}:{last_letter}{first_filter}"
        ws.auto_filter.filterColumn.clear()

    if not df_sleek.empty:
        _write_grand_totals_row(ws, df_sleek, current_row)


# ─── SO COMPANY EXTRACTOR ────────────────────────────────────────────────────
def _so_company(entity: str) -> str:
    e = entity.strip()
    if not e:
        return "(No Entity)"
    parts = re.split(r'\s+-\s+', e, maxsplit=1)
    return parts[0].strip() if parts[0].strip() else e


_GENERIC_SUFFIXES = {
    "advances", "advance", "advantage", "global", "services", "solutions",
    "group", "corp", "corporation", "inc", "ltd", "co", "ph",
    "philippines", "international", "holdings", "enterprises",
}


def _company_root(name: str) -> str:
    tokens = re.sub(r'[^a-z0-9\s]', '', name.lower()).split()
    while len(tokens) > 1 and tokens[-1] in _GENERIC_SUFFIXES:
        tokens = tokens[:-1]
    if not tokens:
        return name.lower()
    if len(tokens) == 1:
        return tokens[0]
    return " ".join(tokens[:2])


def _normalise_so_companies(names: list[str]) -> dict[str, str]:
    from collections import defaultdict

    unique = list(dict.fromkeys(names))

    clusters: dict[str, list[str]] = defaultdict(list)
    for name in unique:
        clusters[_company_root(name)].append(name)

    final_clusters: dict[str, list[str]] = {}
    single_keys = [k for k in clusters if " " not in k]
    multi_keys  = [k for k in clusters if " " in k]

    absorbed: set[str] = set()
    for sk in single_keys:
        target = None
        for mk in multi_keys:
            if mk.startswith(sk + " "):
                target = mk
                break
        if target:
            clusters[target].extend(clusters[sk])
            absorbed.add(sk)

    for key, members in clusters.items():
        if key not in absorbed:
            final_clusters[key] = members

    canon_map: dict[str, str] = {}
    for members in final_clusters.values():
        canonical = min(members, key=lambda n: (len(n), n.lower()))
        for m in members:
            canon_map[m] = canonical

    return canon_map


# ─── SO SHEET WRITER ──────────────────────────────────────────────────────────
def _write_so_sheet(ws, df_so: pd.DataFrame, att_ids: set[str]):
    n_cols      = len(WF_OUTPUT_COLS)
    last_letter = get_column_letter(n_cols)

    df_work = df_so.copy().reset_index(drop=True)
    df_work["_raw_co"]  = df_work["Entity"].apply(lambda e: _so_company(_s(e)))
    canon_map           = _normalise_so_companies(df_work["_raw_co"].tolist())
    df_work["_company"] = df_work["_raw_co"].map(canon_map)

    all_companies   = sorted(df_work["_company"].unique(), key=lambda n: n.lower())
    groups_in_order = []
    for company in all_companies:
        grp = df_work[df_work["_company"] == company].drop(
            columns=["_raw_co", "_company"]
        ).copy()
        if grp.empty:
            continue
        grp = _sort_by_sc(grp)
        if att_ids:
            is_att = grp["Id Number"].isin(att_ids)
            grp = pd.concat([grp[is_att], grp[~is_att]], ignore_index=True)
        groups_in_order.append((company, grp))

    current_row  = 1
    first_filter = None
    col_widths   = [len(h) for h in WF_OUTPUT_COLS]

    section_fill  = PatternFill("solid", fgColor=SECTION_HEADER_BG)
    section_font  = Font(bold=True, color=SECTION_HEADER_COLOR, size=11)
    section_align = Alignment(horizontal="left", vertical="center",
                               indent=1, wrap_text=False)

    for company, grp in groups_in_order:
        sec_cell           = ws.cell(row=current_row, column=1, value=company)
        sec_cell.font      = section_font
        sec_cell.fill      = section_fill
        sec_cell.alignment = section_align
        ws.row_dimensions[current_row].height = 18

        for col_i in range(2, n_cols + 1):
            c       = ws.cell(row=current_row, column=col_i)
            c.fill  = section_fill
            c.value = None
        try:
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row,   end_column=n_cols
            )
        except Exception:
            pass
        current_row += 1

        col_hdr_row = current_row
        for col_i, col_name in enumerate(WF_OUTPUT_COLS, start=1):
            cell           = ws.cell(row=col_hdr_row, column=col_i, value=col_name)
            cell.font      = _make_header_font()
            cell.fill      = _make_header_fill()
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[col_hdr_row].height = 15
        current_row += 1

        if first_filter is None:
            first_filter = col_hdr_row

        for _, row in grp.iterrows():
            rid       = _s(row.get("Id Number", ""))
            is_manual = bool(att_ids) and rid not in att_ids

            for col_i, col_name in enumerate(WF_OUTPUT_COLS, start=1):
                val = row.get(col_name, None)
                if isinstance(val, pd.Series):
                    val = val.iloc[0] if not val.empty else None
                if _isna(val) or val == "":
                    val = None
                if col_name in FINANCIAL_COLS and val is not None:
                    try:
                        val = round(float(val), 2)
                    except (TypeError, ValueError):
                        pass
                cell      = ws.cell(row=current_row, column=col_i, value=val)
                cell.fill = _make_highlight_fill() if is_manual else _make_no_fill()
                if col_name in FINANCIAL_COLS and val is not None:
                    cell.number_format = "#,##0.00"

            for ci, col_name in enumerate(WF_OUTPUT_COLS):
                v = row.get(col_name, None)
                if v is not None and not _isna(v) and str(v).strip():
                    col_widths[ci] = max(col_widths[ci], len(str(v)))

            current_row += 1

        current_row += 1  # spacer between companies

    for ci, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = min(width + 2, 50)

    if first_filter is not None:
        ws.auto_filter.ref = f"A{first_filter}:{last_letter}{first_filter}"
        ws.auto_filter.filterColumn.clear()

    if not df_so.empty:
        _write_grand_totals_row(ws, df_so, current_row)


# ─── PHASE 3: BUILD OUTPUT ────────────────────────────────────────────────────
def build_output(source_wb, df: pd.DataFrame, is_xlsm: bool = False,
                 period_label: str = "") -> bytes:
    gc.collect()

    wb      = source_wb
    att_ids: set[str] = df.attrs.get("att_ids", set())

    # Remove non-essential sheets
    for sname in list(wb.sheetnames):
        if sname not in KEEP_SHEETS:
            try:
                del wb[sname]
            except Exception:
                pass

    gc.collect()

    att_ids_local: set[str] = df.attrs.get("att_ids", set())
    if att_ids_local:
        is_att    = df["Id Number"].isin(att_ids_local)
        df_sorted = pd.concat([df[is_att], df[~is_att]], ignore_index=True)
        df_sorted.attrs.update(df.attrs)
    else:
        df_sorted = df.copy()

    # Write Working File sheet
    ws_wf = wb["Working File"]
    _write_data_sheet(ws_wf, df_sorted, att_ids)

    df_work = df_sorted.copy()
    df_work["_dest"] = df_work["Entity"].apply(lambda e: _entity_sheet(_s(e)))

    # Write entity sheets
    for sname in ENTITY_SHEETS:
        subset = df_work[df_work["_dest"] == sname].drop(columns=["_dest"]).copy()
        subset.attrs.update(df.attrs)

        if att_ids_local and not subset.empty:
            is_att_sub = subset["Id Number"].isin(att_ids_local)
            subset = pd.concat(
                [subset[is_att_sub], subset[~is_att_sub]], ignore_index=True
            )
            subset.attrs.update(df.attrs)

        if sname in wb.sheetnames:
            try:
                del wb[sname]
            except Exception:
                pass
        ws_new = wb.create_sheet(title=sname)

        if sname == "Sleek":
            _write_sleek_sheet(ws_new, subset, att_ids)
        elif sname == "SO":
            _write_so_sheet(ws_new, subset, att_ids)
        else:
            _write_data_sheet_entity(ws_new, subset, att_ids)

        gc.collect()

    # Write Summary sheet
    if "Summary" in wb.sheetnames:
        try:
            del wb["Summary"]
        except Exception:
            pass
    ws_summary = wb.create_sheet(title="Summary")
    _write_summary_sheet(ws_summary, df_sorted, period_label=period_label)

    # ── FIX 5: Safe sheet reordering using wb.worksheets.index() ─────────────
    desired = ["Summary", "Attachment", "Working File"] + ENTITY_SHEETS

    # Drop any stray sheets not in desired
    for sname in list(wb.sheetnames):
        if sname not in desired:
            try:
                del wb[sname]
            except Exception:
                pass

    # Reorder with move_sheet using the corrected .worksheets.index() call
    if hasattr(wb, "move_sheet"):
        for target_idx, sheet_name in enumerate(desired):
            if sheet_name not in wb.sheetnames:
                continue
            try:
                ws_obj      = wb[sheet_name]
                current_idx = wb.worksheets.index(ws_obj)   # ← FIX 5
                offset      = target_idx - current_idx
                if offset != 0:
                    wb.move_sheet(ws_obj, offset=offset)
            except Exception:
                pass

    gc.collect()

    # Save to buffer
    buf = io.BytesIO()
    try:
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
    except Exception as e:
        raise ValueError(f"Failed to save workbook: {e}")
    finally:
        buf.close()
        gc.collect()


# ─── CSS ──────────────────────────────────────────────────────────────────────
def inject_css():
    st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=IBM+Plex+Mono:wght@400;500&family=Lato:wght@300;400;700&display=swap');

        :root {
            --red-900: #7A0000;
            --red-800: #A00000;
            --red-700: #C00000;
            --red-600: #D40000;
            --red-100: #FFF0F0;
            --red-50:  #FFF8F8;
            --ash:     #1C1C1C;
            --charcoal:#2E2E2E;
            --smoke:   #F2EFEC;
            --bone:    #FAF8F6;
            --mist:    #EAE7E4;
            --steel:   #6B6B6B;
            --silver:  #A0A0A0;
            --line:    #E0DDD9;
            --white:   #FFFFFF;
            --green:   #1A7A3C;
            --amber:   #B45309;
        }

        html, body, [class*="css"] {
            font-family: 'Lato', sans-serif;
            color: var(--ash);
        }
        .stApp { background: var(--smoke); }

        .hero {
            background: var(--red-700);
            background-image:
                repeating-linear-gradient(0deg, transparent, transparent 39px,
                    rgba(255,255,255,0.04) 39px, rgba(255,255,255,0.04) 40px),
                repeating-linear-gradient(90deg, transparent, transparent 39px,
                    rgba(255,255,255,0.04) 39px, rgba(255,255,255,0.04) 40px);
            margin: -1rem -1rem 0 -1rem;
            padding: 0 40px; height: 88px;
            display: flex; align-items: center; gap: 20px;
            position: relative; overflow: hidden;
        }
        .hero::after {
            content: ''; position: absolute;
            right: -60px; top: -60px;
            width: 220px; height: 220px; border-radius: 50%;
            background: rgba(255,255,255,0.06); pointer-events: none;
        }
        .hero::before {
            content: ''; position: absolute;
            right: 60px; bottom: -80px;
            width: 160px; height: 160px; border-radius: 50%;
            background: rgba(0,0,0,0.08); pointer-events: none;
        }
        .hero-logo {
            width: 48px; height: 48px;
            background: rgba(255,255,255,0.15);
            border: 2px solid rgba(255,255,255,0.3);
            border-radius: 10px;
            display: flex; align-items: center; justify-content: center;
            font-family: 'Syne', sans-serif;
            font-weight: 800; color: #fff; font-size: 1.4rem;
            flex-shrink: 0; backdrop-filter: blur(4px);
        }
        .hero-logo img { width: 40px; height: 40px; object-fit: contain; border-radius: 6px; }
        .hero-text { flex: 1; }
        .hero-title {
            font-family: 'Syne', sans-serif;
            font-size: 1.15rem; font-weight: 800;
            color: #fff; margin: 0; line-height: 1.1; letter-spacing: -0.3px;
        }
        .hero-sub {
            font-size: 0.68rem; color: rgba(255,255,255,0.65);
            margin: 3px 0 0; letter-spacing: 0.12em;
            text-transform: uppercase; font-weight: 400;
        }
        .hero-pill {
            background: var(--red-900); color: rgba(255,255,255,0.9);
            font-size: 0.6rem; font-weight: 700;
            padding: 5px 12px; border-radius: 20px;
            letter-spacing: 0.1em; text-transform: uppercase;
            border: 1px solid rgba(255,255,255,0.15);
        }

        .sec-label {
            display: flex; align-items: center; gap: 10px; margin: 32px 0 14px;
        }
        .sec-label-line { flex: 1; height: 1px; background: var(--line); }
        .sec-label-text {
            font-family: 'Syne', sans-serif;
            font-size: 0.62rem; font-weight: 700;
            letter-spacing: 0.14em; text-transform: uppercase;
            color: var(--red-700); white-space: nowrap;
        }

        .steps-grid {
            display: grid; grid-template-columns: repeat(3, 1fr);
            gap: 14px; margin: 28px 0 32px;
        }
        .step-card {
            background: var(--white);
            border: 1px solid var(--line);
            border-top: 3px solid var(--red-700);
            border-radius: 0 0 10px 10px;
            padding: 20px 18px 18px; position: relative;
            transition: box-shadow 0.2s;
        }
        .step-card:hover { box-shadow: 0 6px 24px rgba(192,0,0,0.08); }
        .step-num {
            font-family: 'Syne', sans-serif;
            font-size: 2rem; font-weight: 800;
            color: var(--red-100); line-height: 1; margin-bottom: 10px; user-select: none;
        }
        .step-title {
            font-family: 'Syne', sans-serif;
            font-size: 0.82rem; font-weight: 700; color: var(--ash); margin-bottom: 7px;
        }
        .step-desc { font-size: 0.73rem; color: var(--steel); line-height: 1.6; }

        .upload-label {
            font-family: 'Syne', sans-serif;
            font-size: 0.62rem; font-weight: 700;
            letter-spacing: 0.12em; text-transform: uppercase;
            color: var(--red-700); margin-bottom: 8px; display: block;
        }
        [data-testid="stFileUploader"] {
            border: 2px dashed var(--red-700) !important;
            border-radius: 10px !important;
            background: var(--red-50) !important;
            padding: 10px !important;
            transition: background 0.2s !important;
        }
        [data-testid="stFileUploader"]:hover { background: var(--red-100) !important; }

        [data-testid="stBaseButton-primary"],
        .stButton button[kind="primary"] {
            background: var(--red-700) !important; color: #fff !important;
            border: none !important; border-radius: 6px !important;
            font-family: 'Syne', sans-serif !important; font-weight: 700 !important;
            font-size: 0.82rem !important; padding: 11px 28px !important;
            letter-spacing: 0.04em !important; text-transform: uppercase !important;
            transition: background 0.15s !important;
        }
        [data-testid="stBaseButton-primary"]:hover,
        .stButton button[kind="primary"]:hover { background: var(--red-900) !important; }

        .result-banner {
            background: var(--white); border: 1px solid var(--line);
            border-left: 5px solid var(--red-700); border-radius: 8px;
            padding: 18px 22px; display: flex; align-items: center; gap: 16px;
            margin-bottom: 22px;
        }
        .result-icon {
            width: 40px; height: 40px; background: var(--red-700);
            border-radius: 8px; display: flex; align-items: center; justify-content: center;
            font-size: 1.1rem; flex-shrink: 0;
        }
        .result-text {
            font-family: 'Syne', sans-serif;
            font-size: 0.88rem; font-weight: 700; color: var(--ash);
        }
        .result-sub { font-size: 0.71rem; color: var(--steel); margin-top: 3px; }

        .metrics-row { display: grid; gap: 10px; margin: 14px 0; }
        .metrics-row-3 { grid-template-columns: repeat(3, 1fr); }
        .metrics-row-4 { grid-template-columns: repeat(4, 1fr); }
        .metric-card {
            background: var(--white); border: 1px solid var(--line);
            border-bottom: 3px solid var(--red-700);
            border-radius: 8px 8px 0 0; padding: 16px 18px 14px;
        }
        .metric-label {
            font-family: 'Syne', sans-serif;
            font-size: 0.58rem; font-weight: 700;
            text-transform: uppercase; letter-spacing: 0.12em;
            color: var(--red-700); margin-bottom: 8px;
        }
        .metric-value {
            font-family: 'Syne', sans-serif;
            font-size: 1.8rem; font-weight: 800; color: var(--ash); line-height: 1;
        }
        .metric-value.red   { color: var(--red-700); }
        .metric-value.green { color: var(--green); }
        .metric-value.amber { color: var(--amber); }

        [data-testid="stTabs"] [data-baseweb="tab"] {
            font-family: 'Syne', sans-serif !important;
            font-size: 0.74rem !important; font-weight: 700 !important;
            letter-spacing: 0.03em !important;
        }
        [data-testid="stTabs"] [aria-selected="true"] {
            color: var(--red-700) !important;
            border-bottom-color: var(--red-700) !important;
        }

        [data-testid="stDataFrame"] {
            border-radius: 8px !important; overflow: hidden !important;
            border: 1px solid var(--line) !important;
        }

        .stCode, [data-testid="stCode"] {
            font-family: 'IBM Plex Mono', monospace !important;
            font-size: 0.71rem !important; background: var(--ash) !important;
            color: #E8E3DD !important; border: none !important; border-radius: 8px !important;
        }

        [data-testid="stExpander"] {
            border: 1px solid var(--line) !important;
            border-left: 3px solid var(--red-700) !important;
            border-radius: 0 8px 8px 0 !important; background: var(--white) !important;
        }
        [data-testid="stExpanderToggleIcon"] { color: var(--red-700) !important; }

        [data-testid="stTextInput"] input {
            border-radius: 6px !important; border-color: var(--line) !important;
            font-size: 0.82rem !important; background: var(--white) !important;
        }
        [data-testid="stTextInput"] input:focus {
            border-color: var(--red-700) !important;
            box-shadow: 0 0 0 2px rgba(192,0,0,0.12) !important;
        }

        [data-testid="stProgress"] > div > div { background: var(--red-700) !important; }

        hr { border: none !important; border-top: 1px solid var(--line) !important; margin: 28px 0 !important; }

        .footer {
            margin-top: 56px; padding-top: 18px; border-top: 1px solid var(--line);
            text-align: center; font-size: 0.62rem; letter-spacing: 0.1em;
            text-transform: uppercase; color: var(--silver);
        }
        .footer span { color: var(--red-700); font-weight: 700; }

        [data-testid="stCheckbox"] label { font-size: 0.82rem !important; color: var(--ash) !important; }

        .hint-bar {
            background: var(--red-50); border: 1px solid rgba(192,0,0,0.15);
            border-radius: 6px; padding: 10px 14px;
            font-size: 0.71rem; color: var(--red-800); margin-bottom: 18px;
        }
    </style>
    """, unsafe_allow_html=True)


# ─── TOPBAR ───────────────────────────────────────────────────────────────────
def render_topbar():
    logo_html = "<div style='font-family:Syne,sans-serif;font-weight:800;color:#fff;font-size:1.4rem;'>B</div>"
    if os.path.exists(LOGO_PATH):
        import base64
        with open(LOGO_PATH, "rb") as f:
            logo_html = (
                f'<img class="hero-logo-img" '
                f'src="data:image/png;base64,{base64.b64encode(f.read()).decode()}" '
                f'style="width:40px;height:40px;object-fit:contain;border-radius:6px;">'
            )
    st.markdown(f"""
    <div class="hero">
        <div class="hero-logo">{logo_html}</div>
        <div class="hero-text">
            <p class="hero-title">BMG-HMO Automation</p>
            <p class="hero-sub">Billing File Processor · Internal Use Only</p>
        </div>
        <div class="hero-pill">v10.8.2</div>
    </div>
    """, unsafe_allow_html=True)


# ─── HOW IT WORKS ─────────────────────────────────────────────────────────────
def render_how_it_works():
    st.markdown("""
    <div class="steps-grid">
        <div class="step-card">
            <div class="step-num">1</div>
            <div class="step-title">Upload Monthly File</div>
            <div class="step-desc">
                Upload any <strong>.xlsx</strong> or <strong>.xlsm</strong> billing file.
                It must contain an <em>Attachment</em> sheet, a <em>Working File</em> sheet,
                and the word <em>MONTHLY</em> somewhere in the filename or header rows.
            </div>
        </div>
        <div class="step-card">
            <div class="step-num">2</div>
            <div class="step-title">Auto-Sync & Validate</div>
            <div class="step-desc">
                Financial values are refreshed from the Attachment, new employees are
                appended, and no existing row is ever removed.
                Manual / WF-only rows are flagged yellow.
            </div>
        </div>
        <div class="step-card">
            <div class="step-num">3</div>
            <div class="step-title">Download 7-Sheet File</div>
            <div class="step-desc">
                Output contains <strong>Summary</strong> · <strong>Attachment</strong> ·
                <strong>Working File</strong> · <strong>Sleek</strong> · <strong>NYFD</strong> ·
                <strong>BMG Internal</strong> · <strong>SO</strong>.
            </div>
        </div>
    </div>
    <p style="font-size:0.72rem;color:#9CA3AF;margin:-4px 0 20px;">
        ⬛ Only entirely blank rows are excluded. Every employee row is transferred.
    </p>
    """, unsafe_allow_html=True)


# ─── METRICS ──────────────────────────────────────────────────────────────────
def _metric(label: str, value, colour: str = "") -> str:
    cls = f"metric-value {colour}".strip()
    return f"""
    <div class="metric-card">
        <div class="metric-label">{label}</div>
        <div class="{cls}">{value}</div>
    </div>"""


# ─── SYNC STATUS ──────────────────────────────────────────────────────────────
def render_sync_status(df: pd.DataFrame, att_ids: set[str]):
    has_att = bool(att_ids)
    rows = []
    for _, row in df.iterrows():
        rid    = _s(row.get("Id Number", ""))
        entity = _s(row.get("Entity", ""))
        source = (
            ("✅ From Attachment" if rid in att_ids else "⚠️ Manual / WF-only")
            if has_att else "— (no Attachment sheet)"
        )
        rows.append({
            "Id Number":      rid,
            "Name":           _s(row.get("Name",          "")),
            "Entity":         entity,
            "Sheet":          _entity_sheet(entity),
            "Sleek Group":    _sleek_subgroup(entity) if _entity_sheet(entity) == "Sleek" else "—",
            "System Coding":  _s(row.get("System Coding", "")),
            "Medical":        row.get("Medical",              ""),
            "VAT":            row.get("VAT",                  ""),
            "Total Memb Fee": row.get("Total Membership Fee", ""),
            "Source":         source,
        })

    all_df    = pd.DataFrame(rows)
    synced_df = all_df[all_df["Source"] == "✅ From Attachment"].drop(columns=["Source"]).reset_index(drop=True)
    manual_df = all_df[all_df["Source"] == "⚠️ Manual / WF-only"].drop(columns=["Source"]).reset_index(drop=True)

    tab_all, tab_synced, tab_manual, tab_sleek, tab_nyfd, tab_bmg, tab_so = st.tabs([
        f"All  ({len(all_df)})",
        f"✅ Attachment  ({len(synced_df)})",
        f"⚠️ Manual  ({len(manual_df)})",
        f"Sleek  ({len(all_df[all_df['Sheet']=='Sleek'])})",
        f"NYFD  ({len(all_df[all_df['Sheet']=='NYFD'])})",
        f"BMG Internal  ({len(all_df[all_df['Sheet']=='BMG Internal'])})",
        f"SO  ({len(all_df[all_df['Sheet']=='SO'])})",
    ])

    with tab_all:
        st.caption("Every employee across all sheets.")
        st.dataframe(all_df, use_container_width=True, hide_index=True)

    with tab_synced:
        if synced_df.empty:
            st.info("No employees matched any row in the Attachment sheet.")
        else:
            st.caption(f"**{len(synced_df)}** employee(s) matched — financials refreshed.")
            st.dataframe(synced_df, use_container_width=True, hide_index=True)

    with tab_manual:
        if manual_df.empty:
            st.info("No manual / WF-only employees." if has_att
                    else "No Attachment sheet — cannot classify rows.")
        else:
            st.caption(
                f"**{len(manual_df)}** employee(s) absent from Attachment. "
                "Highlighted **yellow** in the downloaded file."
            )
            st.dataframe(manual_df, use_container_width=True, hide_index=True)

    for tab, sname in zip([tab_sleek, tab_nyfd, tab_bmg, tab_so], ENTITY_SHEETS):
        with tab:
            sub = all_df[all_df["Sheet"] == sname].drop(columns=["Sheet"]).reset_index(drop=True)
            if sub.empty:
                st.info(f"No rows assigned to the **{sname}** sheet.")
            elif sname == "Sleek":
                st.caption(f"**{len(sub)}** row(s) → **Sleek** sheet, broken down by country group.")
                for grp_label in SLEEK_SUBGROUPS:
                    grp_df = sub[sub["Sleek Group"] == grp_label].drop(columns=["Sleek Group"]).reset_index(drop=True)
                    if grp_df.empty:
                        continue
                    st.markdown(f"**{grp_label}** — {len(grp_df)} row(s)")
                    st.dataframe(grp_df, use_container_width=True, hide_index=True)
            elif sname == "SO":
                display = sub.drop(columns=["Sleek Group"], errors="ignore").reset_index(drop=True)
                st.caption(f"**{len(display)}** row(s) → **SO** sheet, broken down by company.")
                raw_cos   = display["Entity"].apply(lambda e: _so_company(_s(e))).tolist()
                canon_map = _normalise_so_companies(raw_cos)
                display["_company"] = display["Entity"].apply(
                    lambda e: canon_map.get(_so_company(_s(e)), _so_company(_s(e)))
                )
                for company in sorted(display["_company"].unique(), key=lambda x: x.lower()):
                    co_df = display[display["_company"] == company].drop(columns=["_company"]).reset_index(drop=True)
                    if "System Coding" in co_df.columns:
                        co_df["_sc_rank"] = co_df["System Coding"].apply(lambda v: _sc_sort_key(v)[0])
                        co_df = co_df.sort_values(["_sc_rank", "System Coding"], kind="stable").drop(
                            columns=["_sc_rank"]
                        ).reset_index(drop=True)
                    st.markdown(f"**{company}** — {len(co_df)} row(s)")
                    st.dataframe(co_df, use_container_width=True, hide_index=True)
            else:
                display = sub.drop(columns=["Sleek Group"], errors="ignore").reset_index(drop=True)
                st.caption(f"**{len(display)}** row(s) → **{sname}** sheet.")
                st.dataframe(display, use_container_width=True, hide_index=True)


# ─── ADD EMPLOYEE FORM ────────────────────────────────────────────────────────
def render_add_employee(df: pd.DataFrame) -> pd.DataFrame | None:
    st.caption(
        "Fill the fields below and click **Add Employee**. "
        "The row is appended to both the Working File and the matching entity sheet, "
        "highlighted yellow until it appears in the Attachment."
    )

    with st.form("add_emp", clear_on_submit=True):
        st.markdown("**Identity**")
        c1, c2 = st.columns(2)
        id_num = c1.text_input("Id Number *", placeholder="5060-00999-00-00")
        name   = c2.text_input("Name *",      placeholder="LAST, FIRST M")

        st.markdown("**Financial**")
        f1, f2, f3 = st.columns(3)
        medical = f1.number_input("Medical",              min_value=0.0, step=0.01, format="%.2f", value=0.0)
        vat     = f2.number_input("VAT",                  min_value=0.0, step=0.01, format="%.2f", value=0.0)
        total   = f3.number_input("Total Membership Fee", min_value=0.0, step=0.01, format="%.2f", value=0.0)

        st.markdown("**Classification**")
        entity = st.text_input("Entity", placeholder="e.g. Acme Corp - Advances - Juan")
        st.caption(
            "💡 Entity determines the sheet — "
            "'sleek' → **Sleek**, 'nyfd' → **NYFD**, "
            "'bmg/bmgo internal' → **BMG Internal**, otherwise → **SO**."
        )
        sc = st.selectbox(
            "System Coding",
            [SC_AUTO_LABEL] + list(SC_TO_CAT.keys()),
            index=0,
            help="Leave on the first option to auto-derive from Entity.",
        )
        submitted = st.form_submit_button("Add Employee", type="primary")

    if not submitted:
        return None

    errors = []
    id_num = id_num.strip()
    name   = name.strip()
    if not id_num:
        errors.append("Id Number is required.")
    if not name:
        errors.append("Name is required.")
    if errors:
        for e in errors:
            st.error(e)
        return None

    final_sc   = (derive_system_coding(entity) if entity.strip() else "") \
                  if sc == SC_AUTO_LABEL else sc
    dest_sheet = _entity_sheet(entity)

    nr = {col: "" for col in WF_OUTPUT_COLS}
    nr.update({
        "Id Number":            id_num,
        "Name":                 name.upper(),
        "Entity":               entity.strip(),
        "System Coding":        final_sc,
        "Medical":              float(medical) if medical else None,
        "VAT":                  float(vat)     if vat     else None,
        "Total Membership Fee": float(total)   if total   else None,
    })

    updated = _concat_preserve(df, pd.DataFrame([nr]))
    st.success(
        f"✅ '{nr['Name']}' ({id_num}) added → Working File + **{dest_sheet}** sheet. "
        "Re-download the file below."
    )
    return updated


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="BMG-HMO Automation",
        page_icon="🏥",
        layout="centered",
        initial_sidebar_state="collapsed",
    )
    inject_css()
    render_topbar()

    st.markdown('<div style="height:24px;"></div>', unsafe_allow_html=True)
    render_how_it_works()

    st.markdown(
        '<div class="upload-label">Upload Billing File (.xlsx / .xlsm)</div>',
        unsafe_allow_html=True,
    )
    # FIX 6 — non-empty label, hidden via label_visibility
    uploaded = st.file_uploader(
        "Billing File Upload",
        type=["xlsx", "xlsm"],
        label_visibility="collapsed",
    )

    if not uploaded:
        st.markdown(
            '<div class="footer">BMG-HMO Automation · Internal Use Only · <span>v10.8.2</span></div>',
            unsafe_allow_html=True,
        )
        return

    st.markdown(
        f'<p style="font-size:0.78rem;color:#6B7280;margin:6px 0 16px;">'
        f'<strong style="color:#111827">{uploaded.name}</strong> — ready to process</p>',
        unsafe_allow_html=True,
    )

    # Initialise session state keys
    for key, default in [
        ("df", None), ("out_bytes", None),
        ("file_bytes", None), ("is_xlsm", False),
        ("out_name", ""), ("sync_summary", {}),
        ("period_label", ""), ("proc_logs", []),
    ]:
        if key not in st.session_state:
            st.session_state[key] = default

    if st.button("⚙  Process File", type="primary"):
        prog = st.progress(0, text="Starting…")
        logs: list[str] = []
        wb   = None

        try:
            file_bytes = uploaded.read()
            is_xlsm    = uploaded.name.lower().endswith(".xlsm")

            # FIX 2 — keep buffer open while wb is alive
            buf_read = io.BytesIO(file_bytes)
            wb       = load_workbook(buf_read, keep_vba=is_xlsm)
            # buf_read intentionally left open; wb holds a reference

            prog.progress(10, text="Validating file…")
            logs.append("PHASE 1 · Validation")
            logs.append("─" * 50)
            val = phase1_validate(wb, uploaded.name)
            logs.append(f"  ✓ Monthly billing  : confirmed")
            logs.append(f"  ✓ Attachment sheet : {'found' if val['has_attachment'] else 'MISSING'}")
            logs.append(f"  ✓ Working File     : found")

            period_label = _detect_period_label(wb)
            logs.append(f"  ✓ Period label     : {period_label or '(not detected — using placeholder)'}")

            prog.progress(35, text="Syncing Attachment → Working File…")
            logs.append("")
            logs.append("PHASE 2 · Sync")
            logs.append("─" * 50)
            df, sync_logs = phase2_sync(wb)
            logs.extend(sync_logs)

            prog.progress(65, text="Updating master reference…")
            master       = load_master()
            master, mlog = update_master(df, master)
            save_master(master)
            logs.append("")
            logs.append("MASTER REFERENCE")
            logs.append("─" * 50)
            logs.append(f"  New     : {len(mlog['added'])}")
            logs.append(f"  Updated : {len(mlog['updated'])}")
            logs.append(f"  Missing : {len(mlog['missing'])}")
            logs.append(f"  Total   : {len(master)}")

            prog.progress(80, text="Building output workbook…")

            # FIX 2 — fresh buffer for output wb, kept open
            buf_out = io.BytesIO(file_bytes)
            wb_out  = load_workbook(buf_out, keep_vba=is_xlsm)
            wb_out  = _scrub_workbook(wb_out)
            out_bytes = build_output(wb_out, df, is_xlsm=is_xlsm, period_label=period_label)

            prog.progress(95, text="Finalising…")

            att_ids    = df.attrs.get("att_ids", set())
            entity_col = df["Entity"].apply(lambda e: _entity_sheet(_s(e)))

            st.session_state.df           = df
            st.session_state.out_bytes    = out_bytes
            st.session_state.file_bytes   = file_bytes
            st.session_state.is_xlsm      = is_xlsm
            st.session_state.out_name     = uploaded.name
            st.session_state.period_label = period_label
            st.session_state.proc_logs    = logs
            st.session_state.sync_summary = {
                "total":  len(df),
                "synced": sum(1 for i in df["Id Number"] if i in att_ids),
                "manual": sum(1 for i in df["Id Number"] if i not in att_ids),
                "sleek":  int((entity_col == "Sleek").sum()),
                "nyfd":   int((entity_col == "NYFD").sum()),
                "bmg":    int((entity_col == "BMG Internal").sum()),
                "so":     int((entity_col == "SO").sum()),
            }

            logs.append("")
            logs.append("COMPLETE ✓")
            logs.append("─" * 50)
            logs.append("Sheets: Summary | Attachment | Working File | Sleek | NYFD | BMG Internal | SO")

            # FIX 3 — clear progress bar on success to prevent white-flash rerun loop
            time.sleep(0.4)
            prog.empty()
            gc.collect()

        except ValueError as ve:
            prog.empty()
            st.error(f"Validation Error: {ve}")
        except KeyError as ke:
            prog.empty()
            st.error(f"Sheet/column not found: {ke}")
        except Exception as ex:
            prog.empty()
            st.error(f"Unexpected error: {ex}")
            import traceback
            st.code(traceback.format_exc(), language="python")
        finally:
            # FIX 1 — no wb.close() (openpyxl has none); just let gc handle it
            wb = None
            gc.collect()

    # ── Results panel ──────────────────────────────────────────────────────────
    if st.session_state.df is not None:
        df      = st.session_state.df
        att_ids = df.attrs.get("att_ids", set())
        ss      = st.session_state.sync_summary

        st.markdown('<div style="height:8px;"></div>', unsafe_allow_html=True)

        st.markdown(f"""
        <div class="result-banner">
            <div class="result-icon">✅</div>
            <div>
                <div class="result-text">File processed successfully — ready to download</div>
                <div class="result-sub">
                    {ss.get('total', 0)} records across 7 sheets · Summary sheet included
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.download_button(
            label     = f"⬇  Download {st.session_state.out_name}",
            data      = st.session_state.out_bytes,
            file_name = st.session_state.out_name,
            mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type      = "primary",
        )

        st.markdown('<div class="sec-label">Record Summary</div>', unsafe_allow_html=True)
        st.markdown(f"""
        <div class="metrics-row metrics-row-3">
            {_metric("Total Records",    ss.get('total',  0))}
            {_metric("From Attachment",  ss.get('synced', 0), "green")}
            {_metric("Manual / WF-only", ss.get('manual', 0), "amber")}
        </div>
        <div class="metrics-row metrics-row-4">
            {_metric("→ Sleek",        ss.get('sleek', 0), "red")}
            {_metric("→ NYFD",         ss.get('nyfd',  0))}
            {_metric("→ BMG Internal", ss.get('bmg',   0))}
            {_metric("→ SO",           ss.get('so',    0))}
        </div>
        """, unsafe_allow_html=True)

        with st.expander("🔍  View processing log", expanded=False):
            log_text = "\n".join(st.session_state.get("proc_logs", []))
            st.code(log_text, language=None)

        st.divider()

        st.markdown('<div class="sec-label">Sync Status</div>', unsafe_allow_html=True)
        render_sync_status(df, att_ids)

        st.divider()

        st.markdown('<div class="sec-label">Employee Lookup</div>', unsafe_allow_html=True)
        # FIX 6 — non-empty label, hidden via label_visibility
        search = st.text_input(
            "Employee Lookup",
            placeholder="Search by Id Number or Name…",
            label_visibility="collapsed",
            key="lookup",
        )
        if search:
            mask = (
                df["Id Number"].astype(str).str.contains(search, case=False, na=False) |
                df["Name"].astype(str).str.contains(search, case=False, na=False)
            )
            found = df[mask]
            if not found.empty:
                found_display = found[[c for c in WF_OUTPUT_COLS if c in found.columns]].copy()
                found_display["Sheet"]  = found["Entity"].apply(lambda e: _entity_sheet(_s(e)))
                found_display["Source"] = found["Id Number"].apply(
                    lambda x: "✅ Attachment" if x in att_ids else "⚠️ Manual"
                )
                st.dataframe(found_display, use_container_width=True, hide_index=True)
            else:
                st.info("No records found.")

        st.divider()

        st.markdown(
            '<div class="sec-label">Add New Employee (Optional)</div>',
            unsafe_allow_html=True,
        )
        # FIX 4 — removed value=False to prevent rerun loop
        if st.checkbox("➕  Add a new employee to the Working File", key="show_add_emp"):
            updated_df = render_add_employee(df)
            if updated_df is not None:
                # FIX 2 — keep buffer open while wb2 is alive
                buf_add = io.BytesIO(st.session_state.file_bytes)
                wb2     = load_workbook(buf_add, keep_vba=st.session_state.is_xlsm)
                wb2     = _scrub_workbook(wb2)
                new_bytes = build_output(
                    wb2, updated_df,
                    is_xlsm=st.session_state.is_xlsm,
                    period_label=st.session_state.period_label,
                )
                st.session_state.df        = updated_df
                st.session_state.out_bytes = new_bytes
                gc.collect()
                st.rerun()

    st.markdown(
        '<div class="footer">BMG-HMO Automation · Internal Use Only · <span>v10.8.2</span></div>',
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()
